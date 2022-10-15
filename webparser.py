import requests
from bs4 import BeautifulSoup
import lxml
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side



#pip install requests
#pip install bs4
#pip install lxml

COUNT = 15  ##seller soni
MAIN_URL = 'https://www.ebay.com/sch/i.html?_from=R40&_trksid=p2334524.m570.l1313&_nkw=phone+charger&_sacat=58058&LH_TitleDesc=0&_fsrp=1&rt=nc&_odkw=iphone+charger&_osacat=58058&LH_PrefLoc=1&_fcid=1&_pgn='


def get_products_url(count = COUNT, url = MAIN_URL) -> list:
    h = []
    for p_count in range(1, 10):
        url += str(p_count)
        response = requests.get(url)
        soup = BeautifulSoup(response.text, features='lxml')
        for i in soup.find_all(name  = 'div', attrs={'class' : 's-item__info clearfix'}):
            t = i.find('a')
            try:
                for div in t.next_sibling.next_sibling.find_all('div', {'class':'s-item__detail s-item__detail--primary'}):
                    if div.span['class'][0] == 's-item__location':
                        if div.span.text == 'from United States':
                            #print(div.span.text) #countries
                            if len(h) >= count:
                                return h
                            else:
                                h.append(t.attrs['href'])
            except:
                continue
    return h

def get_sellers_url(products = get_products_url()) -> list:
    sellers = []
    for product_url in products:
        response = requests.get(product_url)
        soup = BeautifulSoup(response.text, features='lxml')
        for seller in soup.find_all('div', attrs={'class':'ux-seller-section__item--seller'}):
            seller_url = seller.find('a')['href']
            #print(seller_url)
            if seller_url not in sellers:
                #print(seller_url) sellers for checking
                sellers.append(seller_url)
    return sellers



def sellers_info(sellers = get_sellers_url()) -> list:
    plan = []
    for seller_url in sellers:
        info = [seller_url,]
        response = requests.get(seller_url)
        soup = BeautifulSoup(response.text, features='lxml')
        for tag in soup.find_all('span', {'class':'info'}):
            since = str(tag.text.strip())
            since_year = int(since.split(' ')[-1])
            if since_year <= 2010:
                info.append(since)
                positive, neutral, negative = soup.find_all('div', {'class':'score'})[0].find('span', {'class', 'num'}).text, soup.find_all('div', {'class':'score'})[1].find('span', {'class', 'num'}).text, soup.find_all('div', {'class':'score'})[2].find('span', {'class', 'num'}).text
                info.append(positive)
                info.append(neutral)
                info.append(negative)
                plan.append(info)
            else:
                continue
        
            
    return plan


def to_excel(plan = sellers_info()):
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(1, 1, 'Seller url')
    cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 24)
    cell = ws.cell(1, 2, 'member since')
    cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 24)
    cell = ws.cell(1, 3, 'Positive score')
    cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 24)
    cell = ws.cell(1, 4, 'Neutral score')
    cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 24)
    cell = ws.cell(1, 5, 'Negative score')
    cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 24)
    cell = ws.cell(1, 6, 'Location')
    cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 24)
    ws.column_dimensions['A'].width = 105.0
    ws.column_dimensions['B'].width = 25.0
    ws.column_dimensions['C'].width = 25.0
    ws.column_dimensions['D'].width = 25.0
    ws.column_dimensions['E'].width = 25.0
    ws.column_dimensions['F'].width = 25.0
    for i in range(len(plan)):
        cell = ws.cell(i+2, 1, plan[i][0])
        cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 19)
        cell = ws.cell(i+2, 2, plan[i][1])
        cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 19)
        cell = ws.cell(i+2, 3, plan[i][2])
        cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 19, color='0000FF00')
        cell = ws.cell(i+2, 4, plan[i][3])
        cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 19, color='FFBB00')
        cell = ws.cell(i+2, 5, plan[i][4])
        cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 19, color='00FF0000')
        cell = ws.cell(i+2, 6, 'United States')
        cell.font = openpyxl.styles.Font(name = 'Iosevka Extended', sz = 19)
        
    wb.save('desktop/tgt/filer.xlsx')

to_excel()





